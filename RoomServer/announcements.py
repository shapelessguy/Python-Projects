import importlib
import os
import sys
import traceback
import json
import pandas as pd
import threading
import time
import telebot
from telebot import types
from datetime import datetime, timedelta
from datetime import date as date_
from openpyxl.styles import Alignment
from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
from utils import LEO_TOKEN, LEO_GROUP_ID, BLAME_LOGS_FILEPATH, WARN_LOGS_FILEPATH, ANNOUNCEMENT_FILEPATH
from utils import MAIN_FOLDER_PATH, MSG_HISTORY_PATH, LAST_IDX_FILEPATH, FINANCE_EXCEL_FILEPATH
from utils import pull, push
from wg import bac
from wg import utils as bac_utils
from bot_ai import reasoner


temp_data = {}
error_msg = "❌ Let's try again shall we?"
signal = None
debug_flag = False
BREAK_TOKEN = '↩️'
hb = None
bot_history = {}


def get_general_metadata():
    emoticons = {a.name: a.emoticons for a in bac.Activities().get_regular()}
    activities = [name for name in emoticons]
    wg_props = [{"name": m.name, "telegram_id": m.telegram_id} for m in bac.WgMembers().get_members()]
    functions = {'get_history': bac_utils.get_history, 'get_plan': bac_utils.get_plan,
                 'get_swaps': bac_utils.get_swaps, 'save_swaps': bac_utils.save_swaps,
                 'get_vacations': bac_utils.get_vacations, 'save_vacations': bac_utils.save_vacations,
                 'get_expenses': bac_utils.get_expenses, 'save_expenses': bac_utils.save_expenses}
    metadata = {'functions': functions, 'wg_props': wg_props, 'activities': activities, 'emoticons': emoticons}
    return metadata


def get_message_md(message, md):
    req_dt = datetime.fromtimestamp(message.date)

    sender_name = None
    for e in md['wg_props']:
        if e['telegram_id'] == message.chat.id:
            sender_name = e['name']
    if sender_name is None:
        return None, None, req_dt
    
    sender_activity = [v for v in md['activities'] if v.lower() == message.text.lower()]
    if len(sender_activity) == 0:
        return None, sender_name, req_dt
    sender_activity = sender_activity[0]

    return sender_activity, sender_name, req_dt


def handle_feedback(req_dt, sender_name, target_activity, feedback_type, save):
    filepath = BLAME_LOGS_FILEPATH if feedback_type == 'blame' else WARN_LOGS_FILEPATH
    with open(filepath, 'r') as file:
        logs = json.load(file)
    current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    previous_week = (current_week - timedelta(7)).strftime("%Y-%m-%d")
    logs[sender_name] = logs.get(sender_name, [])

    entry = {'date': str(previous_week), 'submitted': str(req_dt), feedback_type: target_activity}
    for e in logs[sender_name]:
        if entry['date'] == e['date'] and entry[feedback_type] == e[feedback_type]:
            return datetime.strptime(e['submitted'], "%Y-%m-%d %H:%M:%S"), False

    if save:
        logs[sender_name].append(entry)
        with open(filepath, 'w') as file:
            json.dump(logs, file, indent=2)
    return req_dt, True


def find_target_chat_id(message, req_dt, target_activity, md, feedback_type, sender_name="", continue_chain=True):
    get_history = md['functions']['get_history']
    hist_df = get_history()
    current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    previous_week = (current_week - timedelta(7)).strftime("%Y-%m-%d")
    filtered_row = hist_df[hist_df['Week'] == previous_week]
    target_name = [k for k, v in filtered_row.to_dict().items() if (len(v.values()) > 0 and list(v.values())[0] == target_activity)]
    if len(target_name) == 0:
        new_request(message, f"❌ Nobody to {feedback_type} my friend.", continue_chain)
        return None
    target_name = target_name[0]
    if sender_name == target_name:
        new_request(message, f"❌ Don't be silly! It was you..", continue_chain)
        return None
    target_chat_id = [e['telegram_id'] for e in temp_data[message.chat.id]['md']['wg_props'] if e['name'] == target_name]
    if len(target_chat_id) > 0:
        return target_chat_id[0]
    return None


def blame_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        activities = md['activities']
        sender_activity, sender_name, req_dt = get_message_md(message, md)
        if req_dt.weekday() < 2 and message.data['action'].id == 'blame':
            new_request(message, "❌ Calm down Rocky.. wait until end of Tuesday at least! ✋ And if you didn't do it yet, warn them first!", continue_chain)
            return
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            for i in range(int((len(activities) + 0.99) / 2)):
                if len(activities) > 2*i + 1:
                    markup.row(activities[2*i], activities[2*i + 1])
                else:
                    markup.row(activities[2*i])
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, "Responsible of which activity? (They will never know it was you 😎)", reply_markup=markup)
            if message.data['action'].id == 'blame':
                bh.bot.register_next_step_handler(message, blame_handler2)
            else:
                bh.bot.register_next_step_handler(message, direct_msg_handler1)
    except:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def blame_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_activity, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        if target_activity is None:
            new_request(message, error_msg, continue_chain)
            return

        relevant_date, not_yet_submitted = handle_feedback(req_dt, sender_name, target_activity, 'warn', save=False)
        if not_yet_submitted:
            msg = f"You have to submit a warning at least 1 day before blaming them!"
            new_request(message, msg, continue_chain)
            return
        else:
            hold_on_time = timedelta(hours=24)
            difference = relevant_date + hold_on_time - req_dt
            if difference.total_seconds() > 0:
                hours, remainder = divmod(difference.total_seconds(), 3600)
                minutes = int(remainder // 60)
                msg = f"Wait at least 1 day after the first warning! - {int(hours)}:{'0' if len(str(minutes)) < 2 else ''}{minutes} hours remaining."
                new_request(message, msg, continue_chain)
                return
        relevant_date, saved = handle_feedback(req_dt, sender_name, target_activity, 'blame', save=True)
        if saved:
            # One could also blame himself
            target_chat_id = find_target_chat_id(message, req_dt, target_activity, temp_data[message.chat.id]['md'], 'blame', "", continue_chain)
            if target_chat_id is not None:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
                markup.row(BREAK_TOKEN)
                bh.send_msg(target_chat_id, f"You have been just blamed for {target_activity} my friend.", reply_markup=markup)
            new_request(message, "✅ Thanks for your feedback! 👌", continue_chain)
            return
        else:
            new_request(message, "❌ You have already submitted this feedback 😌", continue_chain)
            return
    except:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def direct_msg_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_activity, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        if target_activity is None:
            new_request(message, error_msg, continue_chain)
            return
        feedback_type = temp_data[message.chat.id]['data']['action'].id  # blame, warn, praise
        target_chat_id = find_target_chat_id(message, req_dt, target_activity, temp_data[message.chat.id]['md'], feedback_type, sender_name, continue_chain)
        if target_chat_id is None:
            return False

        temp_data[message.chat.id]['target_chat_id'] = target_chat_id
        if feedback_type == 'warn':
            handle_feedback(req_dt, sender_name, target_activity, 'warn', save=True)
        current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        pw_format = (current_week - timedelta(days=7)).strftime("%d/%m")
        cw_format = current_week.strftime("%d/%m")
        add = '⚠️ Watch out, you received a warning! 🤨' if feedback_type == 'warn' else '🌟 Someone praised you!'
        temp_data[message.chat.id]['msg_to_send'] = f"For the task {target_activity} (week {pw_format} - {cw_format}):\n{add} - |COMMENT|"
        temp_data[message.chat.id]['msg_to_read'] = f"✅ Alright! {feedback_type.title()} for the task {target_activity} (week {pw_format} - {cw_format}), got u!"
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, "Don't be lazy and leave a comment.. 🙃", reply_markup=markup)
            bh.bot.register_next_step_handler(message, direct_msg_handler2)
    except:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def direct_msg_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_chat_id = temp_data[message.chat.id]['target_chat_id']
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        markup.row(BREAK_TOKEN)
        bh.send_msg(target_chat_id, f"{temp_data[message.chat.id]['msg_to_send'].replace('|COMMENT|', message.text)}", reply_markup=markup)
        new_request(message, f"{temp_data[message.chat.id]['msg_to_read']}", continue_chain)
    except:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def swap_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        _, sender_name, req_dt = get_message_md(message, md)
        if sender_name is None:
            new_request(message, error_msg, continue_chain)
            return
        temp_data[message.chat.id]['sender_name'] = sender_name
        temp_data[message.chat.id]['req_dt'] = req_dt
        activities = md['activities']
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            for i in range(int((len(activities) + 0.99) / 2)):
                if len(activities) > 2*i + 1:
                    markup.row(activities[2*i], activities[2*i + 1])
                else:
                    markup.row(activities[2*i])
            markup.row(BREAK_TOKEN)
            bh.send_msg(
                message.chat.id,
                f"What activity do u wanna swap yours with?",
                reply_markup=markup
            )
            bh.bot.register_next_step_handler(message, swap_handler2)
    except:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)

def swap_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        temp_data[message.chat.id]['target_activity'] = message.text
        if continue_chain:
            calendar, step = DetailedTelegramCalendar().build()
            temp_data[message.chat.id]['calendar_title'] = "Which week are you referring to?"
            temp_data[message.chat.id]['calendar_on_result'] = swap_handler3
            bh.send_msg(
                message.chat.id,
                f"{temp_data[message.chat.id]['calendar_title']} Select {LSTEP[step]}",
                reply_markup=calendar
            )
    except:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)

def swap_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        sender_name = temp_data[message.chat.id]['sender_name']
        req_dt = temp_data[message.chat.id]['req_dt']
        md = temp_data[message.chat.id]['md']
        date = temp_data[message.chat.id]['calendar_result']
        target_activity = temp_data[message.chat.id]['target_activity']
        year, month, day = date.year, date.month, date.day
    
        now = datetime.now()
        now = (now - timedelta(days=now.weekday()))
        date_to_check = datetime(year, month, day)

        current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        if date_to_check < current_week:
            new_request(message, "❌ It's too late now 🙄", continue_chain)
            return

        if date_to_check.weekday() != 0:
            new_request(message, "❌ Date must be a Monday 🤨", continue_chain)
            return
        this_week = now.date() == date_to_check.date()

        swaps = md['functions']['get_swaps']()
        plan_df = md['functions']['get_plan']()

        next_week = (date_to_check + timedelta(days=req_dt.weekday() + 7)).strftime("%d/%m/%Y")
    
        filtered_row = plan_df[plan_df['Week'] == date_to_check.strftime("%Y-%m-%d")]
        date_to_check = date_to_check.strftime("%d/%m/%Y")
        name_to_swap = [k for k, v in filtered_row.to_dict().items() if (len(v.values()) > 0 and list(v.values())[0] == target_activity)]
        try:
            if len(filtered_row) == 0:
                new_request(message, "❌ Is there something even scheduled for that time??", continue_chain)
                return
            sender_activity = [list(v.values())[0] for k, v in filtered_row.to_dict().items() if (len(v.values()) > 0 and k == sender_name)]
            sender_activity = [v for v in md['activities'] if v == sender_activity[0]][0]
            if sender_activity == target_activity:
                new_request(message, "❌ Dude, are you serious? 😡", continue_chain)
                return
        except Exception:
            print(traceback.format_exc())
            name_to_swap = []
        if len(name_to_swap) == 0:
            new_request(message, f"❌ You cannot swap with {target_activity} 😓", continue_chain)
            return
        name_to_swap = name_to_swap[0]
        entry = {'name1': sender_name, 'name2': name_to_swap, 'day': day, 'month': month, 'year': year}
        if entry in swaps['entries']:
            new_request(message, "❌ Dude, you already swapped with this activity 😩", continue_chain)
            return
        swaps['entries'].append(entry)
        md['functions']['save_swaps'](swaps)
        updated_text = '\n🔄 A new plan is now getting generated.. just a moment ah' if this_week else ''
        new_request(message, f"✅ Activity {sender_activity} swapped with {target_activity} for the week {date_to_check} - {next_week}!{updated_text} 🎉", continue_chain)
        if this_week:
            signal['set_announcement'] = {'updated': True}
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def vacations_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        _, sender_name, req_dt = get_message_md(message, md)
        if sender_name is None:
            new_request(message, error_msg, continue_chain)
            return
        current_week = (req_dt - timedelta(days=req_dt.weekday())).date()
        vacations = md['functions']['get_vacations']()
        my_vacations = 'Your vacations:\n'
        found = False
        for v in vacations['entries']:
            date = date_(v['year'], v['month'], v['day'])
            next_week = date + timedelta(weeks=1)
            if sender_name in v['names'] and date >= current_week:
                my_vacations += f'- {date.strftime("%d/%m/%Y")} - {next_week.strftime("%d/%m/%Y")}\n'
                found = True
        my_vacations = my_vacations if found else "You don't have any vacation booked 😲"
        new_request(message, my_vacations, continue_chain)
        return True
    except Exception:
        new_request(message, error_msg, continue_chain)
        return


def vacation_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        if continue_chain:
            calendar, step = DetailedTelegramCalendar().build()
            temp_data[message.chat.id]['target_activity'] = message.text
            temp_data[message.chat.id]['calendar_title'] = "When do you wanna take vacation?"
            temp_data[message.chat.id]['calendar_on_result'] = vacation_handler2
            bh.send_msg(
                message.chat.id,
                f"{temp_data[message.chat.id]['calendar_title']} Select {LSTEP[step]}",
                reply_markup=calendar
            )
    except Exception:
        new_request(message, error_msg, continue_chain)
        return


def vacation_handler2(message, continue_chain=True, unbook_only=False):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        _, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        md = temp_data[message.chat.id]['md']
        if sender_name is None:
            new_request(message, error_msg, continue_chain)
            return
        date = temp_data[message.chat.id]['calendar_result']
        year, month, day = date.year, date.month, date.day

        now = datetime.now()
        now = (now - timedelta(days=now.weekday()))
        date_to_check = datetime(year, month, day)

        current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        if date_to_check < current_week:
            new_request(message, "❌ Dude, you cannot indicate a date in the past! 😡", continue_chain)
            return

        if date_to_check.weekday() != 0:
            new_request(message, "❌ Date must be a Monday! 🤨", continue_chain)
            return
        this_week = now.date() == date_to_check.date()

        vacations = md['functions']['get_vacations']()
        next_week = (date_to_check + timedelta(days=req_dt.weekday() + 7)).strftime("%d/%m/%Y")
        date_to_check = date_to_check.strftime("%d/%m/%Y")

        for v in vacations['entries']:
            if sender_name in v['names'] and v['day'] == day and v['month'] == month and v['year'] == year:
                if continue_chain:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
                    markup.row("NO", "YES")
                    markup.row(BREAK_TOKEN)
                    bh.send_msg(message.chat.id, f"This vacation has already been booked.. do you want to unbook it?", reply_markup=markup)
                    bh.bot.register_next_step_handler(message, vacation_handler3)
                return False
        if unbook_only:
            new_request(message, "❌ Nothing to unbook at this date.. 🤨", continue_chain)
            return
        vacations['entries'].append({'names': [sender_name], 'day': day, 'month': month, 'year': year, 'n_weeks': 1})
        md['functions']['save_vacations'](vacations)
        updated_text = '\n🔄 A new plan is now getting generated.. just a moment please..' if this_week else ''
        new_request(message, f"✅ You booked a vacation for the week {date_to_check} - {next_week}!{updated_text}", continue_chain)
        if this_week:
            signal['set_announcement'] = {'updated': True}
        return
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def vacation_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        if message.text != 'YES':
            new_request(message, f"Alright, then what the hell do you want?? 😤", continue_chain)
            return
        md = temp_data[message.chat.id]['md']
        _, sender_name, req_dt = get_message_md(message, md)
        date = temp_data[message.chat.id]['calendar_result']
        year, month, day = date.year, date.month, date.day

        now = datetime.now()
        now = (now - timedelta(days=now.weekday()))
        date_to_check = datetime(year, month, day)
        this_week = now.date() == date_to_check.date()

        vacations = md['functions']['get_vacations']()
        next_week = (date_to_check + timedelta(days=req_dt.weekday() + 7)).strftime("%d/%m/%Y")
        date_to_check = date_to_check.strftime("%d/%m/%Y")

        for i in range(len(vacations['entries'])):
            v = vacations['entries'][i]
            if sender_name in v['names'] and v['day'] == day and v['month'] == month and v['year'] == year:
                del vacations['entries'][i]
                break
        md['functions']['save_vacations'](vacations)
        updated_text = '\n🔄 A new plan is now getting generated.. just a moment bitte.' if this_week else ''
        new_request(message, f"✅ You unbooked your vacation for the week {date_to_check} - {next_week}.{updated_text}", continue_chain)
        if this_week:
            signal['set_announcement'] = {'updated': True}
        return True
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def expense_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, f"How much did you pay? 🧐", reply_markup=markup)
            bh.bot.register_next_step_handler(message, expense_handler2)
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def expense_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    price = message.text
    try:
        price = [y for x in price.split(',') for y in x.split('.')]
        price = [int(x) for x in price]
        assert 0 < len(price) <= 2
        if len(price) == 1:
            price.append(0)
        price_str = f"{price[0]}.{'0' * (2 - len(str(price[1])))}{price[1]}€"
        price = float(f'{price[0]}.{price[1]}')
        temp_data[message.chat.id]['expense'] = price
        temp_data[message.chat.id]['expense_str'] = price_str

        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, f"For what? 🧐", reply_markup=markup)
            bh.bot.register_next_step_handler(message, expense_handler3)
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def expense_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        _, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        price = temp_data[message.chat.id]['expense']
        price_str = temp_data[message.chat.id]['expense_str']
        reason = message.text
        
        current_week = (req_dt - timedelta(days=req_dt.weekday())).date()
        expenses = temp_data[message.chat.id]['md']['functions']['get_expenses']()
        index = (max([x['index'] for x in expenses['entries']]) + 1) if len(expenses['entries']) > 0 else 0
        expenses['entries'].append({'index': index, 'name': sender_name, 'price': price, 'reason': reason,
                                    'year': current_week.year, 'month': current_week.month, 'day': current_week.day})
        temp_data[message.chat.id]['md']['functions']['save_expenses'](expenses)
        reason_str = f' "{reason}"' if reason != '' else ''
        new_request(message, f'✅ Your expense{reason_str} for {price_str} has been recorded! 🤑', continue_chain)
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def ping_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    new_request(message, f'✅ I am here bro! Chat-id: {message.from_user.id}', continue_chain)


def expenses_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        if continue_chain:
            markup.row("NO", "YES")
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, f"Do you want to continue from last index?", reply_markup=markup)
            bh.bot.register_next_step_handler(message, expenses_handler2)
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


def expenses_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        if message.text == 'YES':
            temp_data[message.chat.id] = {'continue': True}
            expenses_handler3(message)
        elif message.text == 'NO':
            temp_data[message.chat.id] = {'continue': False}
            if continue_chain:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
                markup.row(BREAK_TOKEN)
                bh.send_msg(message.chat.id, f"From which index you want to display the expenses?", reply_markup=markup)
                bh.bot.register_next_step_handler(message, expenses_handler3)
        else:
            new_request(message, "❌ I really don't get it.. 😵", continue_chain)
            return
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)

def expenses_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        in_index = None
        md = get_general_metadata()
        sender_activity, sender_name, req_dt = get_message_md(message, md)
        expenses = md['functions']['get_expenses']()
        max_index = max([x['index'] for x in expenses['entries']]) if len(expenses['entries']) > 0 else -1
        with open(LAST_IDX_FILEPATH, 'r') as file:
            last_idx = {**{sender_name: -1}, **json.load(file)}
        if not temp_data[message.chat.id]['continue']:
            try:
                in_index = int(message.text)
                assert in_index >= 0
                in_index -= 1
            except Exception:
                new_request(message, "❌ Wrong index.", continue_chain)
                return
        else:
            if sender_name in last_idx:
                in_index = last_idx[sender_name]
            else:
                in_index = -1
        
        if in_index >= max_index:
            new_request(message, "❌ No updates here!", continue_chain)
            return

        expenses_list = [
            [e['index'], e['name'], e['price'], e['reason'], date_(day=e['day'], month=e['month'], year=e['year']).strftime("%d/%m/%y")]
            for e in expenses['entries'] if e['index'] > in_index
        ]
        expenses_df = pd.DataFrame(data=expenses_list, columns=['Index', 'Name', 'Price', 'Reason', 'Date'])

        totals_by_name = expenses_df.groupby("Name")["Price"].sum().reset_index()
        totals_by_name.columns = ["Name", "Total Price"]

        with pd.ExcelWriter(FINANCE_EXCEL_FILEPATH, engine="openpyxl") as writer:
            expenses_df.to_excel(writer, sheet_name="Expenses", index=False)
            totals_by_name.to_excel(writer, sheet_name="Totals", index=False)
            sheets = [writer.sheets["Expenses"], writer.sheets["Totals"]]
            
            for sheet in sheets:
                for col in sheet.columns:
                    max_length = 15
                    col_letter = col[0].column_letter
                    sheet.column_dimensions[col_letter].width = max_length
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        new_request(message, f'All the expenses from index {in_index+1}.', continue_chain, document=FINANCE_EXCEL_FILEPATH)
        with open(LAST_IDX_FILEPATH, 'w') as file:
            last_idx[sender_name] = max_index
            json.dump(last_idx, file)
    except Exception:
        print(traceback.format_exc())
        new_request(message, error_msg, continue_chain)


class Action:
    def __init__(self, id, label, handle_request_chain, only_to=None):
        self.id = id
        self.label = label
        self.handle_request_chain = handle_request_chain
        self.only_to = only_to
    
    def start_chain(self, message):
        data = {'action': self}
        message.data = data
        self.handle_request_chain(message)

    def get_label(self, target_id):
        if self.only_to is None or (self.only_to is not None and target_id in self.only_to):
            return self.label


def break_handler(message):
    bh.send_welcome(message)


actions = {
    'blame': Action('blame', '💢 BLAME', blame_handler1),
    'warn': Action('warn', '⚠️ WARN', blame_handler1),
    'praise': Action('praise', '👏 PRAISE', blame_handler1),
    'swap': Action('swap', '🔄 SWAP', swap_handler1),
    'vacations': Action('vacations', '🌴 VACATIONS', vacations_handler1),
    'book': Action('book', '🏖 BOOK VACATION', vacation_handler1),
    'expense': Action('expense', '💰 EXPENSE', expense_handler1),
    'ping': Action('ping', '🛎 PING', ping_handler1, only_to=[807946519]),
    'finance': Action('finance', '📊 Finance', expenses_handler1),
    'break': Action('break', BREAK_TOKEN, break_handler),
}

def get_welcome_markup(target_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    rows = []
    rows.append(['blame', 'warn', 'praise'])
    rows.append(['swap', 'vacations'])
    rows.append(['book', 'expense'])
    rows.append(['ping', 'finance', 'break'])
    rows = [[actions[x].get_label(target_id) for x in row if x is not None] for row in rows]
    rows = [[x for x in row if x is not None] for row in rows]
    rows = [row for row in rows if len(row) > 0]
    for r in rows:
        markup.row(*r)
    return markup


def set_current_id(message, remove_welcome_state=True, continue_chain=True, verbosity=True):
    if not continue_chain:
        return True
    if message.chat.id not in temp_data:
        temp_data[message.chat.id] = {}
    if remove_welcome_state:
        temp_data[message.chat.id]['welcome'] = False
    if 'current_id' in temp_data[message.chat.id] and message.id < temp_data[message.chat.id]['current_id']:
        return False
    temp_data[message.chat.id]['current_id'] = message.id
    if verbosity:
        print(f'Serving {message.chat.id}: msg {message.id}')
    return True


class MyBot:

    def __init__(self):
        self.bot = telebot.TeleBot(LEO_TOKEN)
        self.setup_handlers()
    
    def send_msg(self, *args, **kwargs):
        target_id = str(args[0])
        bot_history[target_id] = bot_history.get(target_id, []) + [{'role': 'assistant', 'message': args[1]}]
        self.bot.send_message(*args, **kwargs)
    
    def send_welcome(self, message):
        if message.chat.id == LEO_GROUP_ID:
            return
        if message.chat.id not in temp_data:
            temp_data[message.chat.id] = {}
        temp_data[message.chat.id]['welcome'] = False
        if message.text == BREAK_TOKEN:
            if not set_current_id(message, False):
                return
        if 'current_id' in temp_data[message.chat.id] and message.id <= temp_data[message.chat.id]['current_id'] and temp_data[message.chat.id]['welcome']:
            return
        print(f'Serving {message.chat.id}: msg {message.id}')

        temp_data[message.chat.id]['welcome'] = True
        self.send_msg(message.chat.id, "How can I be your hero today? 😇", reply_markup=get_welcome_markup(message.chat.id))
        self.bot.register_next_step_handler(message, execute_request)

    def setup_handlers(self):
        @self.bot.message_handler(func=lambda m: True)
        def _(message):
            self.send_welcome(message)

        @self.bot.callback_query_handler(func=DetailedTelegramCalendar.func())
        def cal(c):
            if not set_current_id(c.message):
                return
            if c.message.text == BREAK_TOKEN:
                return
            result, key, step = DetailedTelegramCalendar().process(c.data)
            if not result and key:
                self.bot.edit_message_text(f"{temp_data[c.message.chat.id]['calendar_title']} Select {LSTEP[step]}",
                                    c.message.chat.id,
                                    c.message.message_id,
                                    reply_markup=key)
            elif result:
                temp_data[c.message.chat.id]['calendar_result'] = result
                temp_data[c.message.chat.id]['calendar_on_result'](c.message)


def new_request(message, text, continue_chain=True, document=None):
    target_id = message.chat.id
    if document is not None:
        with open(document, 'rb') as document_:
            bh.bot.send_document(target_id, document_, caption=text, reply_markup=get_welcome_markup(target_id))
    else:
        if not set_current_id(message, verbosity=False):
            return
        bh.send_msg(target_id, text, reply_markup=get_welcome_markup(target_id))
    if continue_chain:
        bh.bot.register_next_step_handler(message, execute_request)


# ------------ DIRECT FUNCTIONS -----------------------------------------------------------------------------------------------------------------


from abc import ABC, abstractmethod


class Function(ABC):
    description = """"""
    arguments = {}
    examples = []

    def __init__(self):
        pass

    @abstractmethod
    def action(self):
        pass


class SendPing(Function):
    description = """
        Send a ping to the chatbot to check if it is still alive.
    """
    arguments = {
    }
    examples = [
        ['send a ping', {}],
        ['can you check if you are alive?', {}],
        ['ping pls', {}],
    ]

    def action(self, message):
        message.data = {'action': actions['ping']}
        ping_handler1(message, continue_chain=False)


wg_activities = [x.name for x in bac.Activities().get_regular()]


class SendWarning(Function):
    description = """
        Send a warning (a telegram message) to the member of the WG which was assigned to the mentioned activity.
        A comment related to the motivation of the warning will accompany the message as well.
    """
    arguments = {
        "activity": f"One of the following {len(wg_activities)} activities ({', '.join(wg_activities)})",
        "comment": "A required comment regarding the warning",
    }
    examples = [
        ['Kitchen is very dirty', 'Do you wanna send a warning to Kitchen?'],
        ['send a warning to Kitchen because it is very dirty', {'activity': 'Kitchen', 'comment': 'Kitchen still dirty.'}],
        ['can you warn management? Nobody disposed the bottles!', {'activity': 'Management', 'comment': 'Bottles not disposed.'}],
        ['execute warning for bathrooms because toilets are shit', {'activity': 'Bathrooms', 'comment': 'Toilets not clean.'}],
        ['execute warning for floor', 'Please provide a comment for this warning.'],
        ['tell bathrooms that they have done an aweful job', {'activity': 'Bathrooms', 'comment': 'Overall bad job.'}],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['warn']}
        blame_handler1(message, continue_chain=False)
        message.text = kwargs['activity']
        if direct_msg_handler1(message, continue_chain=False) is False:
            return
        message.text = kwargs['comment']
        direct_msg_handler2(message, continue_chain=False)
    


class SendPraise(Function):
    description = """
        Send a praise (a telegram message of appreciation) to the member of the WG which was assigned to the mentioned activity.
        A nice comment will accompany the message as well.
    """
    arguments = {
        "activity": f"One of the following {len(wg_activities)} activities ({', '.join(wg_activities)})",
        "comment": "A required comment regarding the praise",
    }
    examples = [
        ['Bathroom did a great job last week', 'Should I praise the responsible then?'],
        ['It s good that management things are done properly', {'activity': 'Management', 'comment': 'Good overall job!'}],
        ['Tell Arman he did a good job at cleaning the bathtub', 'I am not sure, what activity was Arman on? Bathrooms?'],
        ['Praise Floor', 'Please provide a comment for this praise.'],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['praise']}
        blame_handler1(message, continue_chain=False)
        message.text = kwargs['activity']
        if direct_msg_handler1(message, continue_chain=False) is False:
            return
        message.text = kwargs['comment']
        direct_msg_handler2(message, continue_chain=False)


class SendBlame(Function):
    description = """
        Send a blame (a telegram message of disappointment) to the member of the WG which was assigned to the mentioned activity.
        The blame will be recorded in the system and if the subject obtains at least 2 blames in one week, they must redo the activity in the future.
    """
    arguments = {
        "activity": f"One of the following {len(wg_activities)} activities ({', '.join(wg_activities)})"
    }
    examples = [
        ['send a blame to Kitchen because it is very dirty', {'activity': 'Kitchen'}],
        ['Blame Kitchen now!', {'activity': 'Kitchen'}],
        ['The cleaning of Floor has been terrible!', 'Would you like to warn or blame Floor?'],
        ['execute blame for bathrooms because toilets are shit', {'activity': 'Bathrooms'}],
        ['execute blame for Mara', 'Please provide the name of the activity rather than the name of the responsible..'],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['blame']}
        blame_handler1(message, continue_chain=False)
        message.text = kwargs['activity']
        blame_handler2(message, continue_chain=False)


class Swap(Function):
    description = """
        Swap the activity of a given wg member with the activity of another member. The user must indicate what activity they wants to swap their
        activity with and in what date (which must be a Monday in the future with respect to the current date).
    """
    arguments = {
        "activity": f"One of the following {len(wg_activities)} activities ({', '.join(wg_activities)})",
        "date_str": "When the swap has to happen in the form of YYYY-MM-DD",
    }
    examples = [
        ['I want to exchange my task with Mara', "Please provide Mara's activity and the date (remember the date must be referred to a Monday)."],
        ['Swap my task with Management on January', "January doesn't mean anything. The date must be the Monday referred to the week of the swap."],
        ['Swap my task with Floor on the 17th of February 25', {'activity': 'Bathrooms', 'date_str': "2025-02-17"}],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['swap']}
        swap_handler1(message, continue_chain=False)
        message.text = kwargs['activity']
        swap_handler2(message, continue_chain=False)
        temp_data[message.chat.id]['calendar_result'] = datetime.strptime(kwargs['date_str'], "%Y-%m-%d")
        swap_handler3(message, continue_chain=False)


class ShowVacations(Function):
    description = """
        Show the vacations booked by a member of the WG.
    """
    arguments = {}
    examples = [
        ['What are my vacations?', {}],
        ['What are the vacations of Mara', "I have no idea, you could check it by yourself."],
        ['Show all of my vacations on January', 'If you want I can show your next booked vacations..'],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['vacations']}
        vacations_handler1(message, continue_chain=False)


class BookVacation(Function):
    description = """
    Book a vacation in the indicated day (which must be a Monday in the future with respect to the current date).
    The mentioned Monday is identified as the start of the week of vacation.
    The subject is not forced to do any activity during the week they marked as vacation.
    """
    arguments = {
        "date_str": "Date of the vacation in the form of YYYY-MM-DD"
    }
    examples = [
        ['I want take vacation next week', "Then tell me the date (remember the date must be referred to a Monday)."],
        ['I want to book vacations for the 17th of February 25', {'date_str': "2025-02-17"}],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['book']}
        vacation_handler1(message, continue_chain=False)
        temp_data[message.chat.id]['calendar_result'] = datetime.strptime(kwargs['date_str'], "%Y-%m-%d")
        if vacation_handler2(message, continue_chain=False) is False:
            new_request(message, "❌ This vacation has already been booked.", False)


class UnbookVacation(Function):
    description = """
    Unbook a vacation in the indicated day (which must be a Monday in the future with respect to the current date).
    The mentioned Monday is identified as the start of the week of vacation.
    The subject will be subject to an automatically generated activity during the week they marked as non-vacation.
    """
    arguments = {
        "date_str": "Date of the vacation in the form of YYYY-MM-DD"
    }
    examples = [
        ['I want to unbook vacation next week', "Then tell me the date (remember the date must be referred to a Monday)."],
        ["Jokin, I don't want to take any vacation on the 17th of February 25", {'date_str': "2025-02-17"}],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['book']}
        vacation_handler1(message, continue_chain=False)
        temp_data[message.chat.id]['calendar_result'] = datetime.strptime(kwargs['date_str'], "%Y-%m-%d")
        if vacation_handler2(message, continue_chain=False, unbook_only=True) is False:
            message.text = "YES"
            vacation_handler3(message, continue_chain=False)


class CreateExpense(Function):
    description = """
    Create an expense that will be later reimbursed by the WG master. An expense has a price and it is accompanied by a comment which
    indicates what item has been bought by the wg member.
    """
    arguments = {
        "price_str": "Price of the bougth item (or items) as a single float",
        "comment": "A required comment regarding the expense",
    }
    examples = [
        ['I bought plants and chocolate', "Ok, let's do one expense at a time, shall we?"],
        ['I bought plants yesterday', "How much did you pay?"],
        ['I payed 17 euro for the wg', "What did you buy with that money?"],
        ['I payed 4.5 euro for the flowers', {"price_str": "4.50", "comment": "flowers"}],
    ]

    def action(self, message, **kwargs):
        message.data = {'action': actions['expense']}
        expense_handler1(message, continue_chain=False)
        message.text = kwargs['price_str']
        expense_handler2(message, continue_chain=False)
        message.text = kwargs['comment']
        expense_handler3(message, continue_chain=False)


# -----------------------------------------------------------------------------------------------------------------------------------------------


def execute_request(message):
    if message.chat.id == LEO_GROUP_ID:
        return
    target_id = str(message.chat.id)
    bot_history[target_id] = bot_history.get(target_id, []) + [{'role': 'user', 'message': message.text}]
    handled = False
    for action in actions.values():
        if message.text == action.label:
            action.start_chain(message)
            handled = True
            break
    if handled:
        return

    try:
        md = get_general_metadata()
        functions = [f() for f in Function.__subclasses__()]
        reasoner.reason(message, bh, functions, bot_history, md)
        handled = True
    except Exception:
        print(traceback.format_exc())
        bh.send_msg(message.chat.id, "🧠 Sorry I made a mess :S", reply_markup=get_welcome_markup(message.chat.id))
        return

    if handled:
        bh.bot.register_next_step_handler(message, execute_request)


def reload_module(module_name):
    if module_name in sys.modules:
        module = importlib.reload(sys.modules[module_name])
    else:
        module = importlib.import_module(module_name)
    return module


def get_blame(name, now, hist_df, logs):
    filtered_row = hist_df[hist_df['Week'] == now]
    value = [list(v.values())[0] for k, v in filtered_row.to_dict().items() if k == name]
    if len(value) == 0:
        return [], None
    activity = value[0]
    blames = [k for k, v in logs.items() if len([j for j in v if (j['date'] == now and j['blame'] == activity)]) > 0]
    return blames, activity


def set_announcement(updated=False):
    try:
        if not pull(MAIN_FOLDER_PATH, signal):
            return
        with open(BLAME_LOGS_FILEPATH, 'r') as file:
            logs = json.load(file)
        md = get_general_metadata()

        emoticons = md['emoticons']
        now = datetime.now()
        now = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        blame_first_date = (now - timedelta(days=14)).date()
        blame_last_date = (now - timedelta(days=7)).date()
        hist_df = md['functions']['get_history']()
        for e in md['wg_props']:
            name = e['name']
            date = str(blame_first_date)
            blames, blamed_activity = get_blame(name, date, hist_df, logs)
            if len(blames) >= 2:
                blames_json = bac.get_blames()
                blames_json['entries'].append({'name': name, 'date': date})
                bac.save_blames(blames_json)

        intro = '<b>🔄 PLAN UPDATED DURING THIS WEEK</b>\n' if updated else ''
        text, week_schedule = bac.generate_plan()
        text = intro + text
        if not updated:
            text += "\n\nNote: If you want to swap your task with someone else's task, use the SWAP command."
            text += "\nYou can submit an anonymous complaint for one or more tasks of the previous week by sending a BLAME."
            text += "\nBefore blaming someone, please warn them first ⚠️ and then wait until Tuesday evening so that the person has time to recover their delayed duty."
            text += "\nIf a person receives at least 2 blames, they will need to recover the task in the future ☠️."
            text += "\nYou can also send a message of appreciation 🌟 by sending a PRAISE as well."
            text += "\nIf you need vacation for a week, you can send book it though BOOK VACATION."
            text += "\nWhenever you have a WG expense 💰, use EXPENSE and specify price/reason for it."

        if not debug_flag:
            document_ = bac_utils.get_plan_document()
            bh.bot.send_document(LEO_GROUP_ID, document_, caption=text)
        print('Msg sent to LEO6:', text)

        with open(ANNOUNCEMENT_FILEPATH, 'a+') as file:
            file.write(datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + '\n')
        if not push(MAIN_FOLDER_PATH, signal):
            return
        
        for e in md['wg_props']:
            name = e['name']
            blames, blamed_activity = get_blame(name, str(blame_first_date), hist_df, logs)
            telegram_id = e['telegram_id']
            string = '<b>🔄 PLAN UPDATED DURING THIS WEEK</b>\n' if updated else ''
            string += f'Hello {name}, this is the schedule for the next weeks, waiting for you! 🤩\n'
            for week_n in enumerate(week_schedule):
                week_now = (now + timedelta(days=(week_n * 7))).date().strftime("%d/%m")
                week_plus_1 = (now + timedelta(days=((week_n + 1) * 7))).date().strftime("%d/%m")
                pre = 'Week RANGE: ACT\n'
                if week_n == 0:
                    pre = 'This week (RANGE): ACT\n'
                elif week_n == 1:
                    pre = 'Next week (RANGE): ACT\n'
                activity = week_schedule[week_n][name]
                string += pre.replace('RANGE', f'{week_now} - {week_plus_1}').replace('ACT', f' <b>{activity}</b> {emoticons[activity]}')
            if blamed_activity != 'Vacation' and not updated:
                if len(blames) == 0:
                    string += f'\n👍 Congratulations, you have no complaints for the week ' +\
                              f'{blame_first_date.strftime("%d/%m")} - {blame_last_date.strftime("%d/%m")}!'
                else:
                    blamed_activity = f' ({blamed_activity})' if blamed_activity is not None else ''
                    string += f'\n⚠️ You got {len(blames)} complaints in the week {blame_first_date}-{blame_last_date}{blamed_activity}.'
                    if len(blames) >= 2:
                        string += f'\n☠️ Unfortunately you will have to compensate in the next weeks with more tasks.'
            if telegram_id is not None:
                print('Msg sent:', string)
                if name == 'Claudio' or not debug_flag:
                    bh.send_msg(telegram_id, string, parse_mode="HTML")
    except Exception as e:
        print(e)
        print(traceback.format_exc())


def get_week_number(date):
    return date.isocalendar()[1], date.year


def get_date_from_week_id(week_id_):
    d = f"{week_id_[1]}-W{week_id_[0]}"
    return datetime.strptime(d + '-1', "%Y-W%W-%w").date()


def get_last_announcement():
    pull(MAIN_FOLDER_PATH, signal)
    last_announcement = None
    try:
        with open(ANNOUNCEMENT_FILEPATH, 'r') as file:
            last_announcement = datetime.strptime(file.read().split('\n')[-2], "%Y-%m-%d-%H-%M-%S")
    except:
        pass
    if last_announcement is None:
        last_announcement = datetime.strptime("2000-01-01-00-00-00", "%Y-%m-%d-%H-%M-%S")
    return last_announcement, get_week_number(last_announcement)


def spawn_monitoring(signal_):
    global bh, signal, bot_history
    bh = MyBot()
    print('Bot is ready.')
    if os.path.exists(MSG_HISTORY_PATH):
        with open(MSG_HISTORY_PATH, 'r') as file:
            bot_history = json.load(file)
    else:
        bot_history = {}
    signal = signal_
    monitor_thread = threading.Thread(target=monitor_kill, args=(bh, signal))
    monitor_thread.start()
    bh.bot.enable_save_next_step_handlers(delay=2)
    bh.bot.load_next_step_handlers()
    bh.bot.infinity_polling()
    print('Bot killed.')


def monitor_kill(bh, signal):
    index = 0
    while not signal['kill']:
        if 'set_announcement' in signal:
            set_announcement(**signal['set_announcement'])
            del signal['set_announcement']
        index += 1
        if index == 60:
            save_history()
            index = 0
        time.sleep(0.5)
    time.sleep(0.5)
    save_history()
    bh.bot.stop_polling()


def update(last_announcement=None, id_last_announcement=None, forced=False, updated=False):
    global signal
    now = datetime.now()
    announce_at = 8  # hour of update
    masked_now = now - timedelta(hours=(announce_at))
    if forced or get_week_number(masked_now) != id_last_announcement:
        if forced or masked_now > last_announcement:
            print('ANNOUNCEMENTS', now)
            signal['set_announcement'] = {'updated': updated}
            last_announcement = now
            id_last_announcement = get_week_number(now)
    return last_announcement, id_last_announcement


def save_history():
    with open(MSG_HISTORY_PATH, 'w') as file:
        json.dump(bot_history, file, indent=2)


if __name__ == '__main__':
    spawn_monitoring({'kill': False})

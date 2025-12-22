import traceback
import json
import pandas as pd
from utils import LEO_GROUP_ID, BLAME_LOGS_FILEPATH, LAST_IDX_FILEPATH, FINANCE_EXCEL_FILEPATH
from telebot import types
from wg import bac, bac_utils
from datetime import datetime, timedelta
from datetime import date as date_
from openpyxl.styles import Alignment
from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
import telebot
from utils import LEO_TOKEN, LEO_GROUP_ID, BLAME_LOGS_FILEPATH


BREAK_TOKEN = '↩️'
ERROR_MSG = "❌ Let's try again shall we?"
temp_data = {}
bot_history = {}
signal = None


class MyBot:

    def __init__(self):
        self.bot = telebot.TeleBot(LEO_TOKEN)
        self.setup_handlers()
    
    def send_msg(self, *args, **kwargs):
        target_id = str(args[0])
        bot_history[target_id] = bot_history.get(target_id, []) + [{'role': 'assistant', 'message': args[1]}]
        self.bot.send_message(*args, **kwargs)
    
    def send_welcome(self, message):
        if message.chat.id == LEO_GROUP_ID:
            return
        if message.chat.id not in temp_data:
            temp_data[message.chat.id] = {}
        temp_data[message.chat.id]['welcome'] = False
        if message.text == BREAK_TOKEN:
            if not set_current_id(message, False):
                return
        if 'current_id' in temp_data[message.chat.id] and message.id <= temp_data[message.chat.id]['current_id'] and temp_data[message.chat.id]['welcome']:
            return
        print(f'Serving {message.chat.id}: msg {message.id}')

        temp_data[message.chat.id]['welcome'] = True
        self.send_msg(message.chat.id, "How can I be your hero today? 😇", reply_markup=get_welcome_markup(message.chat.id))
        self.bot.register_next_step_handler(message, execute_request)

    def setup_handlers(self):
        @self.bot.message_handler(func=lambda m: True)
        def _(message):
            self.send_welcome(message)

        @self.bot.callback_query_handler(func=DetailedTelegramCalendar.func())
        def cal(c):
            if not set_current_id(c.message):
                return
            if c.message.text == BREAK_TOKEN:
                return
            result, key, step = DetailedTelegramCalendar().process(c.data)
            if not result and key:
                self.bot.edit_message_text(f"{temp_data[c.message.chat.id]['calendar_title']} Select {LSTEP[step]}",
                                    c.message.chat.id,
                                    c.message.message_id,
                                    reply_markup=key)
            elif result:
                temp_data[c.message.chat.id]['calendar_result'] = result
                temp_data[c.message.chat.id]['calendar_on_result'](c.message)


bh = MyBot()


def get_current_time():
    return datetime.now()


def get_welcome_markup(target_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    rows = []
    rows.append(['blame', 'praise', 'tap'])
    rows.append(['expense', 'swap'])
    rows.append(['book', 'vacations'])
    rows.append(['ping', 'finance', 'break'])
    rows = [[actions[x].get_label(target_id) for x in row if x is not None] for row in rows]
    rows = [[x for x in row if x is not None] for row in rows]
    rows = [row for row in rows if len(row) > 0]
    for r in rows:
        markup.row(*r)
    return markup


class Action:
    def __init__(self, id, label, handle_request_chain, only_to=None):
        self.id = id
        self.label = label
        self.handle_request_chain = handle_request_chain
        self.only_to = only_to
    
    def start_chain(self, message):
        data = {'action': self}
        message.data = data
        self.handle_request_chain(message)

    def get_label(self, target_id):
        if self.only_to is None or (self.only_to is not None and target_id in self.only_to):
            return self.label


def new_request(message, text, continue_chain=True, document=None):
    target_id = message.chat.id
    if document is not None:
        with open(document, 'rb') as document_:
            bh.bot.send_document(target_id, document_, caption=text, reply_markup=get_welcome_markup(target_id))
    else:
        if not set_current_id(message, verbosity=False):
            return
        bh.send_msg(target_id, text, reply_markup=get_welcome_markup(target_id))
    if continue_chain:
        bh.bot.register_next_step_handler(message, execute_request)


def execute_request(message):
    if message.chat.id == LEO_GROUP_ID:
        return
    target_id = str(message.chat.id)
    bot_history[target_id] = bot_history.get(target_id, []) + [{'role': 'user', 'message': message.text}]
    handled = False
    for action in actions.values():
        if message.text == action.label:
            action.start_chain(message)
            handled = True
            break
    if handled:
        return

    bh.send_msg(message.chat.id, "🧠 Sorry I don't understand :S", reply_markup=get_welcome_markup(message.chat.id))


def set_current_id(message, remove_welcome_state=True, continue_chain=True, verbosity=True):
    if not continue_chain:
        return True
    if message.chat.id not in temp_data:
        temp_data[message.chat.id] = {}
    if remove_welcome_state:
        temp_data[message.chat.id]['welcome'] = False
    if 'current_id' in temp_data[message.chat.id] and message.id < temp_data[message.chat.id]['current_id']:
        return False
    temp_data[message.chat.id]['current_id'] = message.id
    if verbosity:
        print(f'Serving {message.chat.id}: msg {message.id}')
    return True


def get_general_metadata():
    emoticons = {a.name: a.emoticons for a in bac.Activities().get_all()}
    activities = [name for name in emoticons]
    wg_props = [{"name": m.name, "telegram_id": m.telegram_id} for m in bac.WgMembers(bac.Activities(), bac.Date()).get_members()]
    functions = {'get_history': bac_utils.get_history, 'get_plan': bac_utils.get_plan,
                 'get_swaps': bac_utils.get_swaps, 'save_swaps': bac_utils.save_swaps,
                 'get_vacations': bac_utils.get_vacations, 'save_vacations': bac_utils.save_vacations,
                 'get_expenses': bac_utils.get_expenses, 'save_expenses': bac_utils.save_expenses}
    metadata = {'functions': functions, 'wg_props': wg_props, 'activities': activities, 'emoticons': emoticons}
    return metadata


def get_message_md(message, md):
    req_dt = datetime.fromtimestamp(message.date)

    sender_name = None
    for e in md['wg_props']:
        if e['telegram_id'] == message.chat.id:
            sender_name = e['name']
    if sender_name is None:
        return None, None, req_dt
    
    sender_activity = [v for v in md['activities'] if v.lower() == message.text.lower()]
    if len(sender_activity) == 0:
        return None, sender_name, req_dt
    sender_activity = sender_activity[0]

    return sender_activity, sender_name, req_dt


def handle_blame_feedback(req_dt, sender_name, target_activity, feedback_type, save):
    saved = False
    filepath = BLAME_LOGS_FILEPATH
    with open(filepath, 'r') as file:
        logs = json.load(file)
    current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    previous_week = (current_week - timedelta(7)).strftime("%Y-%m-%d")
    logs[sender_name] = logs.get(sender_name, [])

    entry = {'date': str(previous_week), 'submitted': str(req_dt), feedback_type: target_activity}
    for e in logs[sender_name]:
        if entry['date'] == e['date'] and entry[feedback_type] == e[feedback_type]:
            return saved

    if save:
        logs[sender_name].append(entry)
        with open(filepath, 'w') as file:
            json.dump(logs, file, indent=2)
    saved = True
    return saved


def find_target_chat_id(message, req_dt, target_activity, md, feedback_type, sender_name="", continue_chain=True):
    get_history = md['functions']['get_history']
    hist_df = get_history()
    current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    previous_week = (current_week - timedelta(7 if not feedback_type in ["tap"] else 0)).strftime("%Y-%m-%d")
    filtered_row = hist_df[hist_df['Week'] == previous_week]
    target_name = [k for k, v in filtered_row.to_dict().items() if (len(v.values()) > 0 and list(v.values())[0] == target_activity)]
    if len(target_name) == 0:
        new_request(message, f"❌ Nobody to {feedback_type} my friend.", continue_chain)
        return None, None
    target_name = target_name[0]
    if sender_name == target_name:
        new_request(message, f"❌ Don't be silly! It was you..", continue_chain)
        return None, None
    target_chat_id = [e['telegram_id'] for e in temp_data[message.chat.id]['md']['wg_props'] if e['name'] == target_name]
    if len(target_chat_id) > 0:
        return target_chat_id[0], target_name
    return None, None


def blame_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        activities = [x for x in md['activities'] if x not in ["Vacation", "Blame", "Anarchy"]]
        sender_activity, sender_name, req_dt = get_message_md(message, md)
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            for i in range(int((len(activities) + 1.99) / 2)):
                if len(activities) > 2*i + 1:
                    markup.row(activities[2*i], activities[2*i + 1])
                else:
                    markup.row(activities[2*i])
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, "Responsible of which activity?", reply_markup=markup)
            if message.data['action'].id == 'blame':
                bh.bot.register_next_step_handler(message, blame_handler2)
            elif message.data['action'].id == 'tap':
                bh.bot.register_next_step_handler(message, tap_handler1)
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def blame_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_activity, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        if target_activity is None:
            new_request(message, ERROR_MSG, continue_chain)
            return

        saved = handle_blame_feedback(req_dt, sender_name, target_activity, 'blame', save=True)
        if saved:
            # One could also blame themselves
            target_chat_id, target_name = find_target_chat_id(message, req_dt, target_activity, temp_data[message.chat.id]['md'], 'blame', "", continue_chain)
            if target_chat_id is not None:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
                markup.row(BREAK_TOKEN)
                bh.send_msg(target_chat_id, f"You have been just blamed for {target_activity} my friend.", reply_markup=markup)
            new_request(message, "✅ Thanks for your feedback! 👌", continue_chain)
            return
        else:
            new_request(message, "❌ You have already submitted this feedback 😌", continue_chain)
            return
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def praise_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_activity, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        if target_activity is None:
            new_request(message, ERROR_MSG, continue_chain)
            return
        feedback_type = temp_data[message.chat.id]['data']['action'].id
        target_chat_id, target_name = find_target_chat_id(message, req_dt, target_activity, temp_data[message.chat.id]['md'], feedback_type, sender_name, continue_chain)
        if target_chat_id is None:
            return False
        temp_data[message.chat.id]['target_chat_id'] = target_chat_id
        current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        pw_format = (current_week - timedelta(days=7)).strftime("%d/%m")
        cw_format = current_week.strftime("%d/%m")
        add = '🌟 Someone praised you!'
        temp_data[message.chat.id]['msg_to_send'] = f"For the task {target_activity} (week {pw_format} - {cw_format}):\n{add} - |COMMENT|"
        temp_data[message.chat.id]['msg_to_read'] = f"✅ Alright! {feedback_type.title()} for the task {target_activity} (week {pw_format} - {cw_format}), got u!"
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, "Don't be lazy and leave a comment.. 🙃", reply_markup=markup)
            bh.bot.register_next_step_handler(message, message_handler)
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def tap_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_activity, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        if target_activity is None:
            new_request(message, ERROR_MSG, continue_chain)
            return
        feedback_type = temp_data[message.chat.id]['data']['action'].id  # blame, praise, tap
        target_chat_id, target_name = find_target_chat_id(message, req_dt, target_activity, temp_data[message.chat.id]['md'], feedback_type, sender_name, continue_chain)
        if target_chat_id is None:
            return False

        temp_data[message.chat.id]['target_chat_id'] = target_chat_id
        current_week = (req_dt - timedelta(days=req_dt.weekday()) + timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        pw_format = (current_week - timedelta(days=7)).strftime("%d/%m")
        cw_format = current_week.strftime("%d/%m")
        temp_data[message.chat.id]['msg_to_send'] = f"You have been tapped for the task {target_activity} (week {pw_format} - {cw_format}):\n|COMMENT|"
        temp_data[message.chat.id]['msg_to_read'] = f"✅ Alright! {feedback_type.title()} for the task {target_activity} (week {pw_format} - {cw_format}), got u!"
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, "Leave your precious comment now...", reply_markup=markup)
            bh.bot.register_next_step_handler(message, message_handler)
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def message_handler(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        target_chat_id = temp_data[message.chat.id]['target_chat_id']
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        markup.row(BREAK_TOKEN)
        bh.send_msg(target_chat_id, f"{temp_data[message.chat.id]['msg_to_send'].replace('|COMMENT|', message.text)}", reply_markup=markup)
        new_request(message, f"{temp_data[message.chat.id]['msg_to_read']}", continue_chain)
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def swap_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        _, sender_name, req_dt = get_message_md(message, md)
        if sender_name is None:
            new_request(message, ERROR_MSG, continue_chain)
            return
        temp_data[message.chat.id]['sender_name'] = sender_name
        temp_data[message.chat.id]['req_dt'] = req_dt
        activities = [x for x in md['activities'] if x not in ["Vacation", "Blame", "Anarchy"]]
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            for i in range(int((len(activities) + 1.99) / 2)):
                if len(activities) > 2*i + 1:
                    markup.row(activities[2*i], activities[2*i + 1])
                else:
                    markup.row(activities[2*i])
            markup.row(BREAK_TOKEN)
            bh.send_msg(
                message.chat.id,
                f"What activity do u wanna swap yours with?",
                reply_markup=markup
            )
            bh.bot.register_next_step_handler(message, swap_handler2)
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)

def swap_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        temp_data[message.chat.id]['target_activity'] = message.text
        if continue_chain:
            calendar, step = DetailedTelegramCalendar().build()
            temp_data[message.chat.id]['calendar_title'] = "Which week are you referring to?"
            temp_data[message.chat.id]['calendar_on_result'] = swap_handler3
            bh.send_msg(
                message.chat.id,
                f"{temp_data[message.chat.id]['calendar_title']} Select {LSTEP[step]}",
                reply_markup=calendar
            )
    except:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)

def swap_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        sender_name = temp_data[message.chat.id]['sender_name']
        req_dt = temp_data[message.chat.id]['req_dt']
        md = temp_data[message.chat.id]['md']
        date = temp_data[message.chat.id]['calendar_result']
        target_activity = temp_data[message.chat.id]['target_activity']
        year, month, day = date.year, date.month, date.day
    
        now = get_current_time()
        now = (now - timedelta(days=now.weekday()))
        date_to_check = datetime(year, month, day)

        current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        if date_to_check < current_week:
            new_request(message, "❌ It's too late now 🙄", continue_chain)
            return

        if date_to_check.weekday() != 0:
            new_request(message, "❌ Date must be a Monday 🤨", continue_chain)
            return
        this_week = now.date() == date_to_check.date()

        swaps = md['functions']['get_swaps']()
        plan_df = md['functions']['get_plan']()

        next_week = (date_to_check + timedelta(days=req_dt.weekday() + 7)).strftime("%d/%m/%Y")
    
        filtered_row = plan_df[plan_df['Week'] == date_to_check.strftime("%Y-%m-%d")]
        date_to_check = date_to_check.strftime("%d/%m/%Y")
        name_to_swap = [k for k, v in filtered_row.to_dict().items() if (len(v.values()) > 0 and list(v.values())[0] == target_activity)]
        try:
            if len(filtered_row) == 0:
                new_request(message, "❌ Is there something even scheduled for that time??", continue_chain)
                return
            sender_activity = [list(v.values())[0] for k, v in filtered_row.to_dict().items() if (len(v.values()) > 0 and k == sender_name)]
            sender_activity = [v for v in md['activities'] if v == sender_activity[0]][0]
            if sender_activity == target_activity:
                new_request(message, "❌ Dude, are you serious? 😡", continue_chain)
                return
        except Exception:
            print(traceback.format_exc())
            name_to_swap = []
        if len(name_to_swap) == 0:
            new_request(message, f"❌ You cannot swap with {target_activity} 😓", continue_chain)
            return
        name_to_swap = name_to_swap[0]
        entry = {'name1': sender_name, 'name2': name_to_swap, 'day': day, 'month': month, 'year': year}
        if entry in swaps['entries']:
            new_request(message, "❌ Dude, you already swapped with this activity 😩", continue_chain)
            return
        swaps['entries'].append(entry)
        md['functions']['save_swaps'](swaps)
        updated_text = '\n🔄 A new plan is now getting generated.. just a moment ah' if this_week else ''
        new_request(message, f"✅ Activity {sender_activity} swapped with {target_activity} for the week {date_to_check} - {next_week}!{updated_text} 🎉", continue_chain)
        if this_week:
            signal['set_announcement'] = {'updated': True}
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def vacations_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        _, sender_name, req_dt = get_message_md(message, md)
        if sender_name is None:
            new_request(message, ERROR_MSG, continue_chain)
            return
        current_week = (req_dt - timedelta(days=req_dt.weekday())).date()
        vacations = md['functions']['get_vacations']()
        my_vacations = 'Your vacations:\n'
        found = False
        for v in vacations['entries']:
            date = date_(v['year'], v['month'], v['day'])
            next_week = date + timedelta(weeks=1)
            if sender_name in v['names'] and date >= current_week:
                my_vacations += f'- {date.strftime("%d/%m/%Y")} - {next_week.strftime("%d/%m/%Y")}\n'
                found = True
        my_vacations = my_vacations if found else "You don't have any vacation booked 😲"
        new_request(message, my_vacations, continue_chain)
        return True
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)
        return


def vacation_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        if continue_chain:
            calendar, step = DetailedTelegramCalendar().build()
            temp_data[message.chat.id]['target_activity'] = message.text
            temp_data[message.chat.id]['calendar_title'] = "When do you wanna take vacation?"
            temp_data[message.chat.id]['calendar_on_result'] = vacation_handler2
            bh.send_msg(
                message.chat.id,
                f"{temp_data[message.chat.id]['calendar_title']} Select {LSTEP[step]}",
                reply_markup=calendar
            )
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)
        return


def vacation_handler2(message, continue_chain=True, unbook_only=False):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        _, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        md = temp_data[message.chat.id]['md']
        if sender_name is None:
            new_request(message, ERROR_MSG, continue_chain)
            return
        date = temp_data[message.chat.id]['calendar_result']
        year, month, day = date.year, date.month, date.day

        now = get_current_time()
        now = (now - timedelta(days=now.weekday()))
        date_to_check = datetime(year, month, day)

        current_week = (req_dt - timedelta(days=req_dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        if date_to_check < current_week:
            new_request(message, "❌ Dude, you cannot indicate a date in the past! 😡", continue_chain)
            return

        if date_to_check.weekday() != 0:
            new_request(message, "❌ Date must be a Monday! 🤨", continue_chain)
            return
        this_week = now.date() == date_to_check.date()

        vacations = md['functions']['get_vacations']()
        next_week = (date_to_check + timedelta(days=req_dt.weekday() + 7)).strftime("%d/%m/%Y")
        date_to_check = date_to_check.strftime("%d/%m/%Y")

        for v in vacations['entries']:
            if sender_name in v['names'] and v['day'] == day and v['month'] == month and v['year'] == year:
                if continue_chain:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
                    markup.row("NO", "YES")
                    markup.row(BREAK_TOKEN)
                    bh.send_msg(message.chat.id, f"This vacation has already been booked.. do you want to unbook it?", reply_markup=markup)
                    bh.bot.register_next_step_handler(message, vacation_handler3)
                return False
        if unbook_only:
            new_request(message, "❌ Nothing to unbook at this date.. 🤨", continue_chain)
            return
        vacations['entries'].append({'names': [sender_name], 'day': day, 'month': month, 'year': year, 'n_weeks': 1})
        md['functions']['save_vacations'](vacations)
        updated_text = '\n🔄 A new plan is now getting generated.. just a moment please..' if this_week else ''
        new_request(message, f"✅ You booked a vacation for the week {date_to_check} - {next_week}!{updated_text}", continue_chain)
        if this_week:
            signal['set_announcement'] = {'updated': True}
        return
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def vacation_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        if message.text != 'YES':
            new_request(message, f"Alright, then what the hell do you want?? 😤", continue_chain)
            return
        md = temp_data[message.chat.id]['md']
        _, sender_name, req_dt = get_message_md(message, md)
        date = temp_data[message.chat.id]['calendar_result']
        year, month, day = date.year, date.month, date.day

        now = get_current_time()
        now = (now - timedelta(days=now.weekday()))
        date_to_check = datetime(year, month, day)
        this_week = now.date() == date_to_check.date()

        vacations = md['functions']['get_vacations']()
        next_week = (date_to_check + timedelta(days=req_dt.weekday() + 7)).strftime("%d/%m/%Y")
        date_to_check = date_to_check.strftime("%d/%m/%Y")

        for i in range(len(vacations['entries'])):
            v = vacations['entries'][i]
            if sender_name in v['names'] and v['day'] == day and v['month'] == month and v['year'] == year:
                del vacations['entries'][i]
                break
        md['functions']['save_vacations'](vacations)
        updated_text = '\n🔄 A new plan is now getting generated.. just a moment bitte.' if this_week else ''
        new_request(message, f"✅ You unbooked your vacation for the week {date_to_check} - {next_week}.{updated_text}", continue_chain)
        if this_week:
            signal['set_announcement'] = {'updated': True}
        return True
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def expense_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        md = get_general_metadata()
        temp_data[message.chat.id] = {**temp_data.get(message.chat.id, {}), 'data': message.data, 'md': md}
        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, f"How much did you pay? 🧐", reply_markup=markup)
            bh.bot.register_next_step_handler(message, expense_handler2)
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def expense_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    price = message.text
    try:
        price = [y for x in price.split(',') for y in x.split('.')]
        price = [int(x) for x in price]
        assert 0 < len(price) <= 2
        if len(price) == 1:
            price.append(0)
        price_str = f"{price[0]}.{'0' * (2 - len(str(price[1])))}{price[1]}€"
        price = float(f'{price[0]}.{price[1]}')
        temp_data[message.chat.id]['expense'] = price
        temp_data[message.chat.id]['expense_str'] = price_str

        if continue_chain:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, f"For what? 🧐", reply_markup=markup)
            bh.bot.register_next_step_handler(message, expense_handler3)
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def expense_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        _, sender_name, req_dt = get_message_md(message, temp_data[message.chat.id]['md'])
        price = temp_data[message.chat.id]['expense']
        price_str = temp_data[message.chat.id]['expense_str']
        reason = message.text
        
        current_day = req_dt.date()
        expenses = temp_data[message.chat.id]['md']['functions']['get_expenses']()
        index = (max([x['index'] for x in expenses['entries']]) + 1) if len(expenses['entries']) > 0 else 0
        expenses['entries'].append({'index': index, 'name': sender_name, 'price': price, 'reason': reason,
                                    'year': current_day.year, 'month': current_day.month, 'day': current_day.day})
        temp_data[message.chat.id]['md']['functions']['save_expenses'](expenses)
        reason_str = f' "{reason}"' if reason != '' else ''
        new_request(message, f'✅ Your expense{reason_str} for {price_str} has been recorded! 🤑', continue_chain)
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def ping_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    # document_ = bac_utils.get_plan_document()
    # bh.bot.send_document(message.from_user.id, document_, caption='nada')
    new_request(message, f'✅ I am here bro! Chat-id: {message.from_user.id}', continue_chain)


def expenses_handler1(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
        if continue_chain:
            markup.row("NO", "YES")
            markup.row(BREAK_TOKEN)
            bh.send_msg(message.chat.id, f"Do you want to view last updates only?", reply_markup=markup)
            bh.bot.register_next_step_handler(message, expenses_handler2)
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def expenses_handler2(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        if message.text == 'YES':
            temp_data[message.chat.id] = {'continue': True}
            expenses_handler3(message)
        elif message.text == 'NO':
            temp_data[message.chat.id] = {'continue': False}
            if continue_chain:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
                markup.row(BREAK_TOKEN)
                bh.send_msg(message.chat.id, f"From which index you want to display the expenses?", reply_markup=markup)
                bh.bot.register_next_step_handler(message, expenses_handler3)
        else:
            new_request(message, "❌ I really don't get it.. 😵", continue_chain)
            return
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)

def expenses_handler3(message, continue_chain=True):
    if not set_current_id(message, continue_chain=continue_chain):
        return
    if message.text == BREAK_TOKEN:
        bh.send_welcome(message)
        return
    try:
        in_index = None
        md = get_general_metadata()
        sender_activity, sender_name, req_dt = get_message_md(message, md)
        expenses = md['functions']['get_expenses']()
        max_index = max([x['index'] for x in expenses['entries']]) if len(expenses['entries']) > 0 else -1
        with open(LAST_IDX_FILEPATH, 'r') as file:
            last_idx = {**{sender_name: -1}, **json.load(file)}
        if not temp_data[message.chat.id]['continue']:
            try:
                in_index = int(message.text)
                assert in_index >= 0
                in_index -= 1
            except Exception:
                new_request(message, "❌ Wrong index.", continue_chain)
                return
        else:
            if sender_name in last_idx:
                in_index = last_idx[sender_name]
            else:
                in_index = -1
        
        if in_index >= max_index:
            new_request(message, "❌ No updates here!", continue_chain)
            return

        expenses_list = [
            [e['index'], e['name'], e['price'], e['reason'], date_(day=e['day'], month=e['month'], year=e['year']).strftime("%d/%m/%y")]
            for e in expenses['entries'] if e['index'] > in_index
        ]
        expenses_df = pd.DataFrame(data=expenses_list, columns=['Index', 'Name', 'Price', 'Reason', 'Date'])

        totals_by_name = expenses_df.groupby("Name")["Price"].sum().reset_index()
        totals_by_name.columns = ["Name", "Total Price"]

        with pd.ExcelWriter(FINANCE_EXCEL_FILEPATH, engine="openpyxl") as writer:
            expenses_df.to_excel(writer, sheet_name="Expenses", index=False)
            totals_by_name.to_excel(writer, sheet_name="Totals", index=False)
            sheets = [writer.sheets["Expenses"], writer.sheets["Totals"]]
            
            for sheet in sheets:
                for col in sheet.columns:
                    max_length = 15
                    col_letter = col[0].column_letter
                    sheet.column_dimensions[col_letter].width = max_length
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        new_request(message, f'All the expenses from index {in_index+1}.', continue_chain, document=FINANCE_EXCEL_FILEPATH)
        with open(LAST_IDX_FILEPATH, 'w') as file:
            last_idx[sender_name] = max_index
            json.dump(last_idx, file)
    except Exception:
        print(traceback.format_exc())
        new_request(message, ERROR_MSG, continue_chain)


def break_handler(message):
    bh.send_welcome(message)


actions = {
    'blame': Action('blame', '💢 BLAME', blame_handler1),
    'praise': Action('praise', '👏 PRAISE', blame_handler1),
    'tap': Action('tap', '👆 TAP', blame_handler1),
    'swap': Action('swap', '🔄 SWAP', swap_handler1),
    'vacations': Action('vacations', '🌴 VACATIONS', vacations_handler1),
    'book': Action('book', '🏖 BOOK VACATION', vacation_handler1),
    'expense': Action('expense', '💰 EXPENSE', expense_handler1),
    'ping': Action('ping', '🛎 PING', ping_handler1, only_to=[x.telegram_id for x in bac.WgMembers(bac.Activities(), bac.Date()).get_members() if x.role == "developer"]),
    'finance': Action('finance', '📊 Finance', expenses_handler1),
    'break': Action('break', BREAK_TOKEN, break_handler),
}
